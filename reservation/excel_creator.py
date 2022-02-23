import openpyxl

from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.drawing.image import Image

from reservation.models import Contract, ContractRoom, ContractVilla, ContractTrain, ContractInsurance, ContractTicket, \
    ContractOther, BankAccount, Currency


def excel_last_page(xl_file, xl_info, reservation):
    excel_add_new_page(xl_file, xl_info["page"] + 1)
    xl_sheet = xl_file["umowa"]
    row = (xl_info["page"]) * 63 + 10
    xl_info["page"] += 1
    price_string = ""

    for cur_name, cur_value in reservation.get_all_prices().items():
        if cur_value != 0:
            if price_string != "":
                price_string += f" + "
            price_string += f"{cur_value} {cur_name}"

    xl_sheet[f"B{row}"].value = f"CENA OGÓLNA: {price_string}"
    xl_sheet[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")
    xl_sheet[f"B{row}"].font = Font(name="Calibri Light", size=16, bold=True)
    xl_sheet.merge_cells(f"B{row}:Y{row + 1}")

    row += 4

    for account, value in reservation.get_all_prices().items():
        if value != 0:
            xl_sheet[f"B{row}"].value = f"Płatność na konto {account}:"
            xl_sheet[f"B{row}"].font = Font(name="Calibri Light", size=10, bold=True)
            account_no = BankAccount.objects.get(currency_hash=Currency.objects.get(hash=account)).account_no
            xl_sheet[f"M{row}"].value = " ".join((account_no[:2], account_no[2:6], account_no[6:10], account_no[10:14],
                                                  account_no[14:18], account_no[18:22], account_no[22:26]))
            # xl_sheet[f"B{xl_info['row']}"].alignment = Alignment(horizontal="center", vertical="center")
            xl_sheet[f"M{row}"].font = Font(name="Calibri Light", size=10)
            # xl_sheet.merge_cells(f"B{xl_info['row']}:{xl_info['row'] + 1}")
            row += 1

    row += 2

    xl_sheet[f"B{row}"].value = "Płatność zgodnie z kursem sprzedaży w Santander Bank Polska S.A. w dniu wpłaty."
    xl_sheet[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")
    xl_sheet[f"B{row}"].font = Font(name="Calibri Light", size=10)
    xl_sheet.merge_cells(f"B{row}:Y{row}")

    row += 1

    xl_sheet[f"B{row}"].value = "Kod SWIFT: WBKPPLPPXXX"
    xl_sheet[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")
    xl_sheet[f"B{row}"].font = Font(name="Calibri Light", size=10)
    xl_sheet.merge_cells(f"B{row}:Y{row}")

    row += 3

    xl_sheet[f"B{row}"].value = "Tytuł wpłaty: "
    xl_sheet[f"B{row}"].font = Font(name="Calibri Light", size=10, bold=True)

    xl_sheet[f"F{row}"].value = f"{reservation.owner.first_name} {reservation.owner.last_name} - {reservation.name}"
    xl_sheet[f"B{row}"].font = Font(name="Calibri Light", size=10)

    row = 62 + (xl_info["page"] - 1) * 63

    xl_sheet[f"B{row-15}"].value = "Z warunkami uczestnictwa zapoznałem/łam się i je akceptuję."
    xl_sheet[f"B{row-15}"].font = Font(name="Calibri Light", size=10, bold=True)

    xl_sheet[f"B{row-13}"].value = "Podpis osoby przyjmującej zgłoszenie:"
    xl_sheet[f"B{row-13}"].font = Font(name="Calibri Light", size=10, bold=True)

    xl_sheet[f"P{row-13}"].value = "Podpis zgłaszającego:"
    xl_sheet[f"P{row-13}"].font = Font(name="Calibri Light", size=10, bold=True)

    xl_sheet[f"B{row-6}"].value = reservation.date_of_sign
    xl_sheet.merge_cells(f"B{row-6}:F{row-6}")

    xl_sheet[f"B{row-5}"].value = "................................................................."
    xl_sheet[f"B{row-4}"].value = "Data i czytelny podpis"
    xl_sheet[f"B{row-4}"].font = Font(name="Calibri Light", size=10)

    xl_sheet[f"P{row-5}"].value = "................................................................."
    xl_sheet[f"P{row-4}"].value = "Data i czytelny podpis"
    xl_sheet[f"P{row-4}"].font = Font(name="Calibri Light", size=10)

    return xl_info


def excel_create_price_str(reservation, model):
    price_string = "SUMA: "
    for cur_name, cur_value in reservation.get_prices(model).items():
        if cur_value is not None:
            if price_string != "SUMA: ":
                price_string += f" + "
            price_string += f"{cur_value} {cur_name}"

    return price_string if price_string != "SUMA: " else ""


def excel_new_page_check(xl_file, xl_info, heading):
    last_row = 62 + (xl_info["page"] - 1) * 63
    if heading:
        if xl_info["row"] + 7 > last_row:
            xl_info["page"] += 1
            excel_add_new_page(xl_file, xl_info["page"])
            xl_info["row"] = last_row + 9
    else:
        if xl_info["row"] + 3 > last_row:
            xl_info["page"] += 1
            excel_add_new_page(xl_file, xl_info["page"])
            xl_info["row"] = last_row + 9
    return xl_info


def excel_numerate_plus_logo(xl_file, page_num):
    row = 62
    for page in range(1, page_num + 1):
        xl_file["umowa"].cell(column=13, row=row).value = f'{page} / {page_num}'
        xl_file["umowa"].cell(column=13, row=row).font = Font(size=10, name="Calibri Light")
        img = openpyxl.drawing.image.Image('media/logo.png')
        img.anchor = f"H{row - 61}"
        xl_file["umowa"].add_image(img)
        row += 63


def excel_add_new_page(xl_file, page):
    row = (page - 1) * 63 + 4
    xl_sheet = xl_file["umowa"]

    for column in range(1, 27):
        xl_sheet.cell(row=row - 1, column=column).border = Border(bottom=Side(border_style='hair', color='000000'))
        xl_sheet.cell(row=row + 59, column=column).border = Border(top=Side(border_style='hair', color='000000'))
        xl_sheet[f"M{row + 58}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        xl_sheet.merge_cells(f"M{row + 58}:N{row + 59}")

    for row in range(row, row + 59):
        xl_sheet.cell(row=row, column=1).border = Border(left=Side(border_style='hair', color='000000'))
        xl_sheet.cell(row=row, column=26).border = Border(right=Side(border_style='hair', color='000000'))


def excel_contract_details(xl_file, reservation):
    xl_sheet = xl_file["umowa"]
    xl_sheet["B8"].value = f"UMOWA NR {reservation.name}"
    xl_sheet["B8"].font = Font(name="Calibry Light", size=16, bold=True)
    xl_sheet["B8"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    xl_sheet.merge_cells("B8:Y8")
    xl_sheet["R18"].value = reservation.date_of_contract
    xl_sheet["R21"].value = f"{reservation.get_dates[0]} - {reservation.get_dates[1]}"
    xl_sheet["B28"].value = f"{reservation.owner.first_name} {reservation.owner.last_name}"
    xl_sheet["R28"].value = reservation.owner.date_of_birth
    xl_sheet["B31"].value = f"{reservation.owner.postcode} {reservation.owner.city}"
    xl_sheet["B32"].value = f"{reservation.owner.address}"
    xl_sheet["R31"].value = f"{reservation.owner.phone_number[0:3]}-" \
                            f"{reservation.owner.phone_number[3:6]}-" \
                            f"{reservation.owner.phone_number[6:9]}"


def excel_participants(xl_file, reservation, xl_info):
    xl_sheet = xl_file["umowa"]
    xl_info['row'] = 35
    column = 2
    for index, participant in enumerate(reservation.client.all(), 1):
        xl_sheet.cell(row=xl_info['row'], column=column).value = f"{index}."
        xl_sheet.cell(row=xl_info['row'], column=column + 1).value = f"{participant.first_name} {participant.last_name}"
        xl_sheet.cell(row=xl_info['row'], column=column).font = Font(size=10, name="Calibri Light")
        xl_sheet.cell(row=xl_info['row'], column=column + 1).font = Font(size=10, name="Calibri Light")
        if (xl_info['row'] == 60 or xl_info['row'] == 123) and column == 2:
            xl_info['row'] = 35
            column = 14
        elif (xl_info['row'] == 60 or xl_info['row'] == 123) and column == 14:
            xl_info['page'] += 1
            excel_add_new_page(xl_file, page=xl_info['page'])
            xl_info['row'] = 71
            column = 2
        else:
            xl_info['row'] += 1
    xl_info['row'] += 2
    return xl_info


def excel_insurance(xl_file, xl_info, reservation):
    xl_sheet = xl_file["umowa"]
    thin = Side(border_style="thin", color="000000")
    cell_font = Font(size=8, name="Calibri Light", color="00FFFFFF", bold=True)
    cell_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    cell_fill = PatternFill(start_color='00000000', end_color='00000000', fill_type='solid')
    cell_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    xl_columns = [("B", "E"), ("F", "Y")]
    xl_names = ["Ubezpieczenie", "Wykupione"]
    xl_values = ["Podróżne", "Kosztów rezygnacji"]
    xl_info = excel_new_page_check(xl_file, xl_info, True)
    xl_sheet.cell(column=2, row=xl_info['row']).value = "UBEZPIECZENIA:"
    xl_info['row'] += 2
    xl_info = excel_new_page_check(xl_file, xl_info, True)

    for index, column_pair in enumerate(xl_columns):
        selected_cell = xl_sheet[f"{column_pair[0]}{xl_info['row']}"]
        selected_cell.font = cell_font
        selected_cell.border = cell_border
        selected_cell.alignment = cell_align
        selected_cell.fill = cell_fill
        selected_cell.value = xl_names[index]
        xl_sheet.merge_cells(f"{column_pair[0]}{xl_info['row']}:{column_pair[1]}{xl_info['row'] + 1}")

    for product in [1, 2]:
        xl_info['row'] += product + 1
        xl_info = excel_new_page_check(xl_file, xl_info, False)
        cell_font = Font(size=8, name="Calibri Light")

        xl_sheet[f"B{xl_info['row']}"].font = cell_font
        xl_sheet[f"B{xl_info['row']}"].border = cell_border
        xl_sheet[f"B{xl_info['row']}"].alignment = cell_align
        xl_sheet[f"B{xl_info['row']}"].value = xl_values[product - 1]
        xl_sheet[f"F{xl_info['row']}"].font = cell_font
        xl_sheet[f"F{xl_info['row']}"].border = cell_border
        xl_sheet[f"F{xl_info['row']}"].alignment = cell_align

        if product == 1:
            xl_sheet[f"F{xl_info['row']}"].value = "NIE"
            if ContractInsurance.objects.filter(contract_id=reservation.pk, type=1).exists():
                xl_sheet[f"F{xl_info['row']}"].value = "TAK"
        else:
            xl_sheet[f"F{xl_info['row']}"].value = "NIE"
            if ContractInsurance.objects.filter(contract_id=reservation.pk, type=2).exists():
                xl_sheet[f"F{xl_info['row']}"].value = "TAK"
        xl_sheet.merge_cells(f"B{xl_info['row']}:E{xl_info['row'] + 2}")
        xl_sheet.merge_cells(f"F{xl_info['row']}:Y{xl_info['row'] + 2}")

    xl_info['row'] += 4
    xl_info = excel_new_page_check(xl_file, xl_info, False)
    xl_sheet[f"Y{xl_info['row']}"].value = excel_create_price_str(reservation, ContractInsurance)
    xl_sheet[f"Y{xl_info['row']}"].alignment = Alignment(horizontal="right")
    xl_info['row'] += 2

    return xl_info


def excel_insert_table_data(xl_columns, index, selected_cell, query):
    if xl_columns["names"][index] == "Kierunek":
        try:
            selected_cell.value = f'{query.room.hotel.region.country.name}\n{query.room.hotel.region.name}'
        except:
            selected_cell.value = f'{query.villa.region.country.name}\n{query.villa.region.name}'
    elif xl_columns["names"][index] == "Daty":
        selected_cell.value = f"{query.date_from} -\n{query.date_to}"
    elif xl_columns["names"][index] == "Liczba nocy":
        selected_cell.value = (query.date_to - query.date_from).days
    elif xl_columns["names"][index] == "Hotel":
        selected_cell.value = query.room.hotel.name
    elif xl_columns["names"][index] == "Zakwaterowanie":
        selected_cell.value = query.room.name
    elif xl_columns["names"][index] == "Wyżywienie":
        selected_cell.value = query.meal_plan.short_name
    elif xl_columns["names"][index] == "Willa":
        selected_cell.value = query.villa.name
    elif xl_columns["names"][index] == "Trasa":
        selected_cell.value = f"{query.train.dest_from} -\n{query.train.dest_to}"
    elif xl_columns["names"][index] == "Pociąg":
        selected_cell.value = query.train.name
    elif xl_columns["names"][index] == "Kabina":
        selected_cell.value = query.cabin_name
    elif xl_columns["names"][index] == "Linia":
        selected_cell.value = query.airline.name
    elif xl_columns["names"][index] == "Klasa":
        selected_cell.value = query.get_ticket_class_display()
    elif xl_columns["names"][index] == "Liczba biletów":
        selected_cell.value = query.quantity
    elif xl_columns["names"][index] == "Szczegóły":
        selected_cell.value = query.ticket_details
    elif xl_columns["names"][index] == "Usługa":
        selected_cell.value = query.name
    elif xl_columns["names"][index] == "Opis":
        selected_cell.value = query.description


def excel_create_table(xl_file, xl_columns, xl_info, model, reservation):
    xl_sheet = xl_file["umowa"]
    thin = Side(border_style="thin", color="000000")
    cell_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    xl_info = excel_new_page_check(xl_file, xl_info, True)

    # print(xl_columns['title'])
    xl_sheet.cell(column=2, row=xl_info['row']).value = xl_columns['title']
    xl_info['row'] += 2
    xl_info = excel_new_page_check(xl_file, xl_info, True)

    for index, column_pair in enumerate(xl_columns['tags']):
        xl_cell = xl_sheet[f"{column_pair[0]}{xl_info['row']}"]
        xl_cell.font = Font(size=8, name='Calibri Light', color='00FFFFFF', bold=True)
        xl_cell.border = cell_border
        xl_cell.fill = PatternFill(start_color='00000000', end_color='00000000', fill_type='solid')
        xl_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        xl_cell.value = xl_columns["names"][index]
        xl_sheet.merge_cells(f"{column_pair[0]}{xl_info['row']}:{column_pair[1]}{xl_info['row'] + 1}")
    xl_info['row'] += 2
    for query in model.objects.filter(contract_id=reservation.pk):
        for index, column_pair in enumerate(xl_columns['tags']):
            xl_info = excel_new_page_check(xl_file, xl_info, False)
            xl_cell = xl_sheet[f"{column_pair[0]}{xl_info['row']}"]
            xl_cell.font = Font(size=8, name="Calibri Light")
            xl_cell.border = cell_border
            xl_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            selected_cell = xl_sheet[f"{column_pair[0]}{xl_info['row']}"]
            excel_insert_table_data(xl_columns, index, selected_cell, query)
            xl_sheet.merge_cells(f"{column_pair[0]}{xl_info['row']}:{column_pair[1]}{xl_info['row'] + 2}")
        xl_info['row'] += 3
    xl_info['row'] += 1
    xl_sheet[f"Y{xl_info['row']}"].value = excel_create_price_str(reservation, model)
    xl_sheet[f"Y{xl_info['row']}"].alignment = Alignment(horizontal="right")
    xl_info['row'] += 3
    return xl_info


def contract_to_excel(id):
    column_data = [
        {
            "model": ContractRoom,
            "tags": [("B", "E"), ("F", "I"), ("J", "K"), ("L", "Q"), ("R", "V"), ("W", "Y")],
            "names": ['Kierunek', 'Daty', 'Liczba nocy', 'Hotel', 'Zakwaterowanie', 'Wyżywienie'],
            "title": "REZERWACJE HOTELOWE:"
        },
        {
            "model": ContractVilla,
            "tags": [("B", "E"), ("F", "I"), ("J", "K"), ("L", "V"), ("W", "Y")],
            "names": ['Kierunek', 'Daty', 'Liczba nocy', 'Willa', 'Wyżywienie'],
            "title": "WYNAJEM WILLI:"
        },
        {
            "model": ContractTrain,
            "tags": [("B", "E"), ("F", "I"), ("J", "K"), ("L", "Q"), ("R", "V"), ("W", "Y")],
            "names": ['Trasa', 'Daty', 'Liczba nocy', 'Pociąg', 'Kabina', 'Wyżywienie'],
            "title": "BILETY LOTNICZE:"
        },
        {
            "model": ContractTicket,
            "tags": [("B", "E"), ("F", "H"), ("I", "J"), ("K", "Y")],
            "names": ['Linia', 'Klasa', 'Liczba biletów', 'Szczegóły'],
            "title": "PRZEJAZDY POCIĄGAMI:"
        },
        {
            "model": ContractOther,
            "tags": [("B", "E"), ("F", "I"), ("J", "Y")],
            "names": ['Usługa', 'Daty', 'Opis'],
            "title": "USŁUGI DODATKOWE:"
        }
    ]

    xl_info = {'row': 13, 'page': 1}
    reservation = Contract.objects.get(pk=id)
    xl_file = openpyxl.load_workbook("media/instyle.xlsx")
    excel_contract_details(xl_file, reservation)
    xl_info = excel_participants(xl_file, reservation, xl_info)

    for index, model_data in enumerate(column_data):
        model = model_data["model"]
        if model.objects.filter(contract_id=id):
            xl_columns = column_data[index]
            xl_info = excel_create_table(xl_file, xl_columns, xl_info, model, reservation)
            xl_columns.clear()

    xl_info = excel_insurance(xl_file, xl_info, reservation)
    xl_info = excel_last_page(xl_file, xl_info, reservation)
    excel_numerate_plus_logo(xl_file, page_num=xl_info['page'])

    xl_file.save(f"media/{reservation.name.replace('/','_')}.xlsx")
